#!/usr/bin/env python3
"""
inputs.json -> sell/buy-side 스타일 financial model (.xlsx).

구조(베이스 = 정통 3-statement):
  Dashboard / Income Statement / Balance Sheet / Cash Flow / Ratios
  + Segment build / Revenue bridge / Margin bridge / Leading indicators
  + Peer benchmark(선택) / Assumptions & Sources

핵심 원칙: Bloomberg 템플릿에 값이 있는 라인아이템(IS/BS/CF/Ratios)은
하나도 빼지 말고 그대로 옮긴다(inputs.statements). 그 위에 분석 탭을 add 한다.
bridge·growth-delta는 라이브 Excel 수식.

사용: python build_model.py --inputs inputs.json --out model.xlsx
"""
import json, argparse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

H=Font(bold=True,color="FFFFFF",size=11); HF=PatternFill("solid",fgColor="1F3864")
SUB=Font(bold=True,color="1F3864"); SUBF=PatternFill("solid",fgColor="D6E0F0")
LBL=Font(bold=True); ITAL=Font(italic=True,color="808080",size=9)
PCT='0.0%'; NUM='#,##0.0'; NUM0='#,##0'; XFMT='0.0"x"'
def colL(i): return get_column_letter(i)
def vals(it): return it.get("values",[]) if isinstance(it,dict) else (it or [])
def fmt_of(it):
    f=(it.get("fmt") or "").lower() if isinstance(it,dict) else ""
    return {"pct":PCT,"x":XFMT,"num0":NUM0}.get(f,NUM)
def src_comment(cell,it):
    if isinstance(it,dict):
        s=it.get("source",""); t=it.get("tier","")
        if s or t: cell.comment=Comment(f"source: {s}\ntier: {t}","model")

def header_row(ws,row,label,periods,c0=1):
    c=ws.cell(row=row,column=c0,value=label); c.font=H; c.fill=HF
    for j,p in enumerate(periods):
        x=ws.cell(row=row,column=c0+1+j,value=p); x.font=H; x.fill=HF; x.alignment=Alignment(horizontal="center")

def write_line(ws,row,it,periods,n,default_fmt=NUM,comment=True):
    """statement line dict {label,values,fmt,bold,source,tier,indent}. 값 없으면 섹션헤더."""
    label=it.get("label","") if isinstance(it,dict) else str(it)
    indent="   "*int(it.get("indent",0)) if isinstance(it,dict) else ""
    v=vals(it); has=any(x not in ("",None) for x in v)
    cell=ws.cell(row=row,column=1,value=indent+label)
    if (isinstance(it,dict) and it.get("bold")) or not has: cell.font=LBL
    fmt=fmt_of(it) if isinstance(it,dict) and it.get("fmt") else default_fmt
    for j in range(n):
        c=ws.cell(row=row,column=2+j)
        c.value=v[j] if j<len(v) and v[j] not in ("",None) else None
        c.number_format=fmt
        if j==0 and comment: src_comment(c,it)

def build_statement(wb,title,block,periods,n,default_fmt=NUM):
    """전체 라인아이템을 그대로 옮긴 statement 시트. key->row 맵 반환."""
    ws=wb.create_sheet(title)
    ws.cell(row=1,column=1,value=title).font=SUB
    r=3; header_row(ws,r,"",periods); r+=1
    keymap={}
    for it in block.get("lines",[]):
        write_line(ws,r,it,periods,n,default_fmt)
        k=it.get("key") if isinstance(it,dict) else None
        if k: keymap[k]=r
        r+=1
    ws.column_dimensions['A'].width=34
    for j in range(n): ws.column_dimensions[colL(2+j)].width=12
    return ws,keymap


def build_quarterly(wb, q, annual_periods, unit, is_name, KEY):
    """분기 시트: 분기 컬럼 + FY컬럼(=SUM(분기) 라이브 링크). flow=합, stock=Q4, ratio=재계산.
    연간 IS와 FY tie-out 체크 포함."""
    qs=q["quarters"]; groups=q.get("fy_groups",{})
    q2fy={}; 
    for fy,qq in groups.items():
        for x in qq: q2fy[x]=fy
    plan=[]; col=2; fycols={}
    for i,ql in enumerate(qs):
        plan.append(("q",ql,col,i)); col+=1
        fy=q2fy.get(ql)
        last=(i==len(qs)-1) or (q2fy.get(qs[i+1])!=fy)
        if fy and last:
            qcols=[pp[2] for pp in plan if pp[0]=="q" and q2fy.get(pp[1])==fy]
            plan.append(("fy",fy,col,qcols)); fycols[fy]=col; col+=1
    ws=wb.create_sheet("Quarterly")
    ws.cell(row=1,column=1,value=f"Quarterly model [{unit}] — 분기 + FY(=SUM 분기, 라이브 링크)").font=SUB
    hr=3
    for kind,label,c,*rest in plan:
        cell=ws.cell(row=hr,column=c,value=("FY"+str(label)[-2:] if kind=="fy" else label))
        cell.font=H; cell.fill=(PatternFill("solid",fgColor="2E5496") if kind=="fy" else HF); cell.alignment=Alignment(horizontal="center")
    r=hr+1; keyrows={}
    for it in q["lines"]:
        label=it.get("label",""); indent="   "*int(it.get("indent",0)); typ=it.get("type","flow"); v=it.get("values",[])
        has=any(x not in ("",None) for x in v)
        lc=ws.cell(row=r,column=1,value=indent+label)
        if it.get("bold") or not has: lc.font=LBL
        for kind,lbl,c,*rest in plan:
            cell=ws.cell(row=r,column=c)
            if kind=="q":
                qi=rest[0]; cell.value=v[qi] if qi<len(v) and v[qi] not in ("",None) else None
            else:
                qcols=rest[0]
                if typ=="flow": cell.value="="+"+".join(f"{colL(cc)}{r}" for cc in qcols)
                elif typ=="stock": cell.value=f"={colL(qcols[-1])}{r}"
            cell.number_format=fmt_of(it)
            if kind=="fy": cell.font=LBL
        k=it.get("key")
        if k: keyrows[k]=r
        r+=1
    if "OperatingProfit" in keyrows and "Revenue" in keyrows:
        ws.cell(row=r,column=1,value="Operating margin %").font=ITAL
        for kind,lbl,c,*rest in plan:
            ws.cell(row=r,column=c,value=f"=IF({colL(c)}{keyrows['Revenue']}=0,\"\",{colL(c)}{keyrows['OperatingProfit']}/{colL(c)}{keyrows['Revenue']})").number_format=PCT
        r+=1
    # FY tie-out vs annual IS (Revenue)
    if "Revenue" in keyrows and "Revenue" in KEY:
        ws.cell(row=r,column=1,value="FY Rev vs 연간 IS (tie-out)").font=ITAL
        for fy,c in fycols.items():
            ann=None
            for j,p in enumerate(annual_periods):
                if str(p).startswith(str(fy)): ann=2+j; break
            if ann:
                ws.cell(row=r,column=c,value=f"={colL(c)}{keyrows['Revenue']}-'{is_name}'!{colL(ann)}{KEY['Revenue']}").number_format=NUM
        r+=1
    ws.column_dimensions['A'].width=24
    for kind,lbl,c,*rest in plan: ws.column_dimensions[colL(c)].width=9
    return fycols

def build(inp,out):
    meta=inp.get("meta",{}); periods=meta.get("periods",[]); n=len(periods)
    freq=meta.get("frequency","A"); lag=4 if str(freq).upper().startswith("Q") else 1
    unit=meta.get("unit","")
    wb=openpyxl.Workbook(); wb.remove(wb.active)

    st=inp.get("statements",{})
    IS_NAME="Income Statement"; KEY={}
    if st.get("income_statement",{}).get("lines"):
        # ---- full 3-statement from Bloomberg/inputs (nothing dropped) ----
        _,KEY=build_statement(wb,IS_NAME,st["income_statement"],periods,n)
        if st.get("balance_sheet",{}).get("lines"):
            build_statement(wb,"Balance Sheet",st["balance_sheet"],periods,n)
        if st.get("cash_flow",{}).get("lines"):
            build_statement(wb,"Cash Flow",st["cash_flow"],periods,n)
        if st.get("ratios",{}).get("lines"):
            build_statement(wb,"Ratios",st["ratios"],periods,n,default_fmt=PCT)
    else:
        # ---- fallback: simple historicals (back-compat) ----
        h=inp.get("historicals",{}); ws=wb.create_sheet(IS_NAME)
        ws.cell(row=1,column=1,value=f"{meta.get('name','')} — Income Statement [{unit}]").font=SUB
        r=3; header_row(ws,r,"P&L",periods); r+=1
        for key,label in [("Revenue","Revenue"),("COGS","COGS"),("GrossProfit","Gross Profit"),
                          ("OpEx","OpEx"),("OperatingProfit","Operating Profit"),("NetIncome","Net Income")]:
            it=h.get(key,{}); it=dict(it,label=label) if isinstance(it,dict) else {"label":label,"values":it}
            write_line(ws,r,it,periods,n); KEY[key]=r; r+=1
        ws.column_dimensions['A'].width=30
        for j in range(n): ws.column_dimensions[colL(2+j)].width=12

    def isc(j): return f"'{IS_NAME}'!{colL(2+j)}"
    # ---- derived rows appended to IS: GM%, OPM%, YoY, growth delta ----
    isws=wb[IS_NAME]; r=isws.max_row+2
    isws.cell(row=r,column=1,value="— Derived (analytics) —").font=SUB; isws.cell(row=r,column=1).fill=SUBF; r+=1
    def derived(label,formula_fn,fmt=PCT):
        nonlocal r
        isws.cell(row=r,column=1,value=label).font=ITAL
        for j in range(n):
            f=formula_fn(j)
            if f: isws.cell(row=r,column=2+j,value=f).number_format=fmt
        rr=r; r+=1; return rr
    gm=opm=grow=gd=None
    if "GrossProfit" in KEY and "Revenue" in KEY:
        gm=derived("Gross margin %",lambda j:f"=IF({isc(j)}{KEY['Revenue']}=0,\"\",{isc(j)}{KEY['GrossProfit']}/{isc(j)}{KEY['Revenue']})")
    if "OperatingProfit" in KEY and "Revenue" in KEY:
        opm=derived("Operating margin %",lambda j:f"=IF({isc(j)}{KEY['Revenue']}=0,\"\",{isc(j)}{KEY['OperatingProfit']}/{isc(j)}{KEY['Revenue']})")
    if "Revenue" in KEY:
        grow=derived("Revenue YoY %",lambda j:(f"=IF({isc(j-lag)}{KEY['Revenue']}=0,\"\",{isc(j)}{KEY['Revenue']}/{isc(j-lag)}{KEY['Revenue']}-1)" if j>=lag else None))
        gd=derived("Growth delta (Δ YoY,%p)",lambda j:(f"=IF(OR({colL(2+j)}{grow}=\"\",{colL(2+j-1)}{grow}=\"\"),\"\",{colL(2+j)}{grow}-{colL(2+j-1)}{grow})" if j>=lag+1 else None))

    # ============ Quarterly (분기 + FY 링크) ============
    q=inp.get("quarterly",{})
    if q.get("quarters") and q.get("lines"):
        build_quarterly(wb, q, periods, unit, IS_NAME, KEY)

    # ============ Segment build ============
    members=inp.get("segments",{}).get("members",[])
    seg_rev_rows=[]; tot_row=None
    if members:
        ss=wb.create_sheet("Segment build")
        ss.cell(row=1,column=1,value=f"Segment build — axis: {inp.get('segments',{}).get('axis_primary','')} [{unit}]").font=SUB
        r=3; header_row(ss,r,"Revenue by segment",periods); r+=1
        for m in members:
            write_line(ss,r,dict(m.get("revenue",{}),label=m.get("name","?")),periods,n); seg_rev_rows.append((m.get("name"),r)); r+=1
        tot_row=r; ss.cell(row=r,column=1,value="Total (segment sum)").font=LBL
        for j in range(n):
            if seg_rev_rows:
                ss.cell(row=r,column=2+j,value=f"=SUM({colL(2+j)}{seg_rev_rows[0][1]}:{colL(2+j)}{seg_rev_rows[-1][1]})").number_format=NUM
        r+=1
        if "Revenue" in KEY:
            ss.cell(row=r,column=1,value="vs IS Revenue (tie-out)").font=ITAL
            for j in range(n):
                ss.cell(row=r,column=2+j,value=f"=IF({colL(2+j)}{tot_row}=0,\"\",{colL(2+j)}{tot_row}-{isc(j)}{KEY['Revenue']})").number_format=NUM
            r+=1
        r+=1; header_row(ss,r,"Segment mix %",periods); r+=1
        for name,rr in seg_rev_rows:
            ss.cell(row=r,column=1,value=name).font=ITAL
            for j in range(n):
                ss.cell(row=r,column=2+j,value=f"=IF({colL(2+j)}{tot_row}=0,\"\",{colL(2+j)}{rr}/{colL(2+j)}{tot_row})").number_format=PCT
            r+=1
        if any(vals(m.get("op",{})) for m in members):
            r+=1; header_row(ss,r,"Operating profit by segment",periods); r+=1
            for m in members:
                write_line(ss,r,dict(m.get("op",{}),label=m.get("name","?")),periods,n); r+=1
        ss.column_dimensions['A'].width=26
        for j in range(n): ss.column_dimensions[colL(2+j)].width=12

    # ============ Revenue bridge ============
    if seg_rev_rows:
        rb=wb.create_sheet("Revenue bridge")
        rb.cell(row=1,column=1,value="Revenue bridge — segment 기여도 (YoY, %p of group growth)").font=SUB
        rb.cell(row=2,column=1,value="각 segment의 전사 YoY 성장률 기여. 합 = 전사 YoY 성장률.").font=ITAL
        r=4; header_row(rb,r,"Contribution to YoY growth (%p)",periods); r+=1
        contrib=[]
        for name,rvr in seg_rev_rows:
            rb.cell(row=r,column=1,value=name).font=LBL
            for j in range(n):
                if j>=lag:
                    cur=f"'Segment build'!{colL(2+j)}{rvr}"; pr=f"'Segment build'!{colL(2+j-lag)}{rvr}"; pt=f"'Segment build'!{colL(2+j-lag)}{tot_row}"
                    rb.cell(row=r,column=2+j,value=(f"=IF(OR(NOT(ISNUMBER({cur})),NOT(ISNUMBER({pr})),{pt}=0),\"\",({cur}-{pr})/{pt})")).number_format=PCT
            contrib.append(r); r+=1
        rb.cell(row=r,column=1,value="Total YoY growth (check)").font=LBL
        for j in range(n):
            if j>=lag and contrib:
                rb.cell(row=r,column=2+j,value=f"=SUM({colL(2+j)}{contrib[0]}:{colL(2+j)}{contrib[-1]})").number_format=PCT
        r+=2
        vp=[m for m in members if vals(m.get("volume",{})) and vals(m.get("asp",{}))]
        if vp:
            rb.cell(row=r,column=1,value=f"Volume / Price 분해 [{unit} 단위 주의: volume×ASP native]").font=SUB; r+=1
            for m in vp:
                header_row(rb,r,m.get("name"),periods); r+=1
                write_line(rb,r,dict(m["volume"],label="  Volume"),periods,n); vr=r; r+=1
                write_line(rb,r,dict(m["asp"],label="  ASP"),periods,n); ar=r; r+=1
                rb.cell(row=r,column=1,value="  Volume effect").font=ITAL
                for j in range(n):
                    if j>=lag: rb.cell(row=r,column=2+j,value=f"=({colL(2+j)}{vr}-{colL(2+j-lag)}{vr})*{colL(2+j-lag)}{ar}").number_format=NUM
                r+=1
                rb.cell(row=r,column=1,value="  Price effect").font=ITAL
                for j in range(n):
                    if j>=lag: rb.cell(row=r,column=2+j,value=f"=({colL(2+j)}{ar}-{colL(2+j-lag)}{ar})*{colL(2+j-lag)}{vr}").number_format=NUM
                r+=2
        rb.column_dimensions['A'].width=28
        for j in range(n): rb.column_dimensions[colL(2+j)].width=12

    # ============ Margin bridge ============
    if {"GrossProfit","Revenue","OperatingProfit","OpEx"}.issubset(KEY) and gm:
        mb=wb.create_sheet("Margin bridge")
        mb.cell(row=1,column=1,value=f"Margin bridge — ΔGP / ΔOP 분해 [{unit}]").font=SUB
        r=3; header_row(mb,r,"Gross profit bridge (YoY)",periods); r+=1
        mb.cell(row=r,column=1,value="ΔGP (actual)").font=LBL
        for j in range(n):
            if j>=lag: mb.cell(row=r,column=2+j,value=f"={isc(j)}{KEY['GrossProfit']}-{isc(j-lag)}{KEY['GrossProfit']}").number_format=NUM
        dgp=r; r+=1
        mb.cell(row=r,column=1,value="  Volume/Revenue effect").font=ITAL
        for j in range(n):
            if j>=lag: mb.cell(row=r,column=2+j,value=f"=({isc(j)}{KEY['Revenue']}-{isc(j-lag)}{KEY['Revenue']})*{isc(j-lag)}{gm}").number_format=NUM
        r+=1
        mb.cell(row=r,column=1,value="  Margin (GM%) effect").font=ITAL
        for j in range(n):
            if j>=lag: mb.cell(row=r,column=2+j,value=f"={isc(j)}{KEY['Revenue']}*({isc(j)}{gm}-{isc(j-lag)}{gm})").number_format=NUM
        r+=2; header_row(mb,r,"Operating profit bridge (YoY)",periods); r+=1
        mb.cell(row=r,column=1,value="ΔOP (actual)").font=LBL
        for j in range(n):
            if j>=lag: mb.cell(row=r,column=2+j,value=f"={isc(j)}{KEY['OperatingProfit']}-{isc(j-lag)}{KEY['OperatingProfit']}").number_format=NUM
        r+=1
        mb.cell(row=r,column=1,value="  GP effect").font=ITAL
        for j in range(n):
            if j>=lag: mb.cell(row=r,column=2+j,value=f"={colL(2+j)}{dgp}").number_format=NUM
        r+=1
        mb.cell(row=r,column=1,value="  Opex leverage effect").font=ITAL
        for j in range(n):
            if j>=lag: mb.cell(row=r,column=2+j,value=f"=-({isc(j)}{KEY['OpEx']}-{isc(j-lag)}{KEY['OpEx']})").number_format=NUM
        r+=2
        mb.cell(row=r,column=1,value="GM% drivers:").font=LBL; mb.cell(row=r,column=2,value=inp.get("margin_bridge",{}).get("gm_drivers_note","")); r+=1
        mb.cell(row=r,column=1,value="Opex leverage:").font=LBL; mb.cell(row=r,column=2,value=inp.get("margin_bridge",{}).get("opex_leverage_note",""))
        mb.column_dimensions['A'].width=28
        for j in range(n): mb.column_dimensions[colL(2+j)].width=12

    # ============ Leading indicators ============
    li_in=inp.get("leading_indicators",{})
    if any(vals(li_in.get(k,{})) for k in li_in if not k.startswith("_")):
        li=wb.create_sheet("Leading indicators")
        li.cell(row=1,column=1,value="Leading indicators — 선행지표 (지속성)").font=SUB
        r=3; header_row(li,r,"Indicator",periods); r+=1
        names={"backlog":"수주잔고","book_to_bill":"Book-to-bill","export_volume":"수출 물량","export_value":"수출 금액",
               "capex":"Capex","capacity":"생산능력","utilization":"가동률","inventory":"재고"}
        bl=None
        for k,lab in names.items():
            it=li_in.get(k,{})
            if not vals(it): continue
            fmt=PCT if k=="utilization" else (NUM if k=="book_to_bill" else NUM0)
            write_line(li,r,dict(it,label=lab),periods,n,default_fmt=fmt)
            if k=="backlog": bl=r
            r+=1
        if bl:
            li.cell(row=r,column=1,value="수주잔고 YoY %").font=ITAL
            for j in range(n):
                if j>=lag: li.cell(row=r,column=2+j,value=f"=IF({colL(2+j-lag)}{bl}=0,\"\",{colL(2+j)}{bl}/{colL(2+j-lag)}{bl}-1)").number_format=PCT
            r+=1
        li.column_dimensions['A'].width=24
        for j in range(n): li.column_dimensions[colL(2+j)].width=12

    # ============ Peer benchmark (optional) ============
    peers=inp.get("peers",{})
    if peers.get("columns") and peers.get("rows"):
        pb=wb.create_sheet("Peer benchmark")
        pb.cell(row=1,column=1,value="Peer Benchmark").font=Font(bold=True,size=13,color="1F3864")
        pb.cell(row=2,column=1,value=f"출처: {peers.get('source','')} (tier {peers.get('tier','')}) — 정성/정량 혼합, 모델 하드넘버 아님").font=ITAL
        cols=peers["columns"]; hr=4
        c=pb.cell(row=hr,column=1,value="구분"); c.font=H; c.fill=HF
        for j,cn in enumerate(cols):
            c=pb.cell(row=hr,column=2+j,value=cn); c.font=H; c.fill=HF; c.alignment=Alignment(wrap_text=True,vertical="top")
        r=hr+1
        for row in peers["rows"]:
            pb.cell(row=r,column=1,value=row[0] if row else "").font=SUB
            for j in range(len(cols)):
                cc=pb.cell(row=r,column=2+j,value=row[1+j] if 1+j<len(row) else ""); cc.alignment=Alignment(wrap_text=True,vertical="top")
            r+=1
        pb.column_dimensions['A'].width=22
        for j in range(len(cols)): pb.column_dimensions[colL(2+j)].width=38

    # ============ Dashboard (front) ============
    db=wb.create_sheet("Dashboard"); wb.move_sheet("Dashboard",-(len(wb.sheetnames)-1))
    db.cell(row=1,column=1,value=f"{meta.get('name','')} ({meta.get('ticker','')}) — Growth Delta Dashboard").font=Font(bold=True,size=14,color="1F3864")
    db.cell(row=2,column=1,value="핵심 질문: 성장의 delta가 커지는가, 지속 가능한가?").font=ITAL
    r=4; header_row(db,r,"성장 추세",periods); r+=1
    if "Revenue" in KEY:
        db.cell(row=r,column=1,value="Revenue").font=LBL
        for j in range(n): db.cell(row=r,column=2+j,value=f"={isc(j)}{KEY['Revenue']}").number_format=NUM
        r+=1
    if grow:
        db.cell(row=r,column=1,value="YoY growth %").font=LBL
        for j in range(n): db.cell(row=r,column=2+j,value=f"=IF('{IS_NAME}'!{colL(2+j)}{grow}=\"\",\"\",'{IS_NAME}'!{colL(2+j)}{grow})").number_format=PCT
        r+=1
    if gd:
        db.cell(row=r,column=1,value="Growth delta (가속,%p)").font=Font(bold=True,color="C00000")
        for j in range(n): db.cell(row=r,column=2+j,value=f"=IF('{IS_NAME}'!{colL(2+j)}{gd}=\"\",\"\",'{IS_NAME}'!{colL(2+j)}{gd})").number_format=PCT
        r+=1
    if opm:
        db.cell(row=r,column=1,value="Operating margin %").font=LBL
        for j in range(n): db.cell(row=r,column=2+j,value=f"={isc(j)}{opm}").number_format=PCT
        r+=1
    r+=1
    db.cell(row=r,column=1,value="지속성 신호: Leading indicators 탭 / 성장의 질: Revenue bridge × Margin bridge").font=ITAL
    db.column_dimensions['A'].width=30
    for j in range(n): db.column_dimensions[colL(2+j)].width=12

    # ============ Assumptions & Sources ============
    asn=wb.create_sheet("Assumptions & Sources")
    asn.cell(row=1,column=1,value="Assumptions & Sources — 입력 출처·tier").font=SUB
    for i,hh in enumerate(["항목","값","출처","tier","비고"]):
        c=asn.cell(row=3,column=1+i,value=hh); c.font=H; c.fill=HF
    r=4
    def add(label,it,note=""):
        nonlocal r
        if not isinstance(it,dict): return
        v=it.get("values",it.get("value"))
        asn.cell(row=r,column=1,value=label); asn.cell(row=r,column=2,value=str(v) if v not in (None,[],"") else "")
        asn.cell(row=r,column=3,value=it.get("source","")); asn.cell(row=r,column=4,value=it.get("tier","")); asn.cell(row=r,column=5,value=note); r+=1
    # statement-level source (한 줄 요약)
    for nm,blk in [("Income Statement",st.get("income_statement")),("Balance Sheet",st.get("balance_sheet")),
                   ("Cash Flow",st.get("cash_flow")),("Ratios",st.get("ratios"))]:
        if blk: add(f"{nm} (전 라인)", {"value":blk.get("source",""),"source":blk.get("source",""),"tier":blk.get("tier","")},"Bloomberg 등 원본 보존")
    for k,it in inp.get("historicals",{}).items(): add(f"Historicals/{k}",it)
    for m in members:
        for fld in ("revenue","op","volume","asp"):
            if fld in m: add(f"Segment/{m.get('name')}/{fld}",m[fld])
    for k,it in li_in.items():
        if not k.startswith("_"): add(f"Leading/{k}",it)
    for a in inp.get("assumptions",[]):
        asn.cell(row=r,column=1,value=a.get("item","")); asn.cell(row=r,column=2,value=str(a.get("value","")))
        asn.cell(row=r,column=3,value=a.get("source","")); asn.cell(row=r,column=4,value=a.get("tier","")); asn.cell(row=r,column=5,value=a.get("note","")); r+=1
    r+=1; asn.cell(row=r,column=1,value="가설 (Qualitative, tier5 — 숫자 미반영)").font=SUB; r+=1
    for q in inp.get("qualitative",[]):
        asn.cell(row=r,column=1,value=q.get("thesis","")); asn.cell(row=r,column=3,value=q.get("source",""))
        asn.cell(row=r,column=4,value=q.get("tier",5)); asn.cell(row=r,column=5,value=("1차소스 확인" if q.get("primary_source_confirmed") else "미확인")); r+=1
    r+=1; asn.cell(row=r,column=1,value="Data gaps").font=SUB; r+=1
    for g in inp.get("data_gaps",[]): asn.cell(row=r,column=1,value=f"• {g}"); r+=1
    for i,w in enumerate([34,18,30,6,30]): asn.column_dimensions[colL(1+i)].width=w

    wb.save(out); print(f"wrote {out}\nsheets: {wb.sheetnames}")

if __name__=="__main__":
    ap=argparse.ArgumentParser()
    ap.add_argument("--inputs",required=True); ap.add_argument("--out",required=True)
    ap.add_argument("--template",default=None)
    a=ap.parse_args()
    build(json.load(open(a.inputs)),a.out)
