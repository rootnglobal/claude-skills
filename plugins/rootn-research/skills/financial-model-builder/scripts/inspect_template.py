#!/usr/bin/env python3
"""
Bloomberg FA 템플릿(_Q.xlsx / _A.xlsx) 구조를 JSON으로 덤프.
사용: python inspect_template.py <template.xlsx> [--out structure.json]
- 시트 목록, 각 시트 크기
- Segments 탭의 type/group_by/members/기간 컬럼
- IS 라인아이템 + 캐시값 존재 여부
캐시값(data_only=True)을 우선 읽어 add-in 비활성(#NAME?) 상황을 처리한다.
"""
import sys, json, argparse, re
import openpyxl

def find_segments_block(ws):
    out = {"found": False}
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 400)):
        for c in row:
            if isinstance(c.value, str) and c.value.strip() in ("Segments", "Segment"):
                # scan nearby cells for header fields
                meta = {}
                for r in ws.iter_rows(min_row=c.row, max_row=min(c.row+8, ws.max_row)):
                    cells = [x.value for x in r]
                    for i, v in enumerate(cells):
                        if isinstance(v, str) and v.strip() in (
                            "Ticker","Name","Acc Standard","Currency","Type","Group By","Period","Level","Consolidated"):
                            nxt = cells[i+1] if i+1 < len(cells) else None
                            if nxt not in (None, ""):
                                meta[v.strip()] = nxt
                out = {"found": True, "anchor_row": c.row, "meta": meta}
                return out
    return out

def period_headers(ws):
    pat = re.compile(r"(FY|FQ|Q)[0-9]?\s?20[0-9]{2}|20[0-9]{2}[ ]?Q[1-4]|FY20[0-9]{2}")
    hits = []
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 60)):
        for c in row:
            if isinstance(c.value, str) and pat.search(c.value):
                hits.append({"cell": c.coordinate, "value": c.value})
    return hits[:60]


def extract_statement(ws, max_scan=400):
    """Best-effort: find period header row, then emit {periods, lines:[{label,values}]}.
    Bloomberg 시트는 지저분하므로 라벨(텍스트)+숫자행만 추출. Claude가 검증/정렬한다."""
    import re
    per_pat = re.compile(r"(FY|FQ)\s?\d{0,2}\s?20\d{2}|20\d{2}[ ]?Q[1-4]|20\d{2}A|20\d{2}E|FY20\d{2}")
    hdr_row=None; per_cols=[]
    for ri,row in enumerate(ws.iter_rows(min_row=1,max_row=min(ws.max_row,60)),1):
        hits=[(c.column,str(c.value)) for c in row if isinstance(c.value,str) and per_pat.search(c.value)]
        if len(hits)>=2:
            hdr_row=ri; per_cols=[h[0] for h in hits]; periods=[h[1].strip() for h in hits]; break
    if not hdr_row: return None
    lines=[]
    for row in ws.iter_rows(min_row=hdr_row+1,max_row=min(ws.max_row,max_scan)):
        # label = first text cell in cols 1..per_cols[0]-1
        label=None
        for c in row:
            if c.column>=per_cols[0]: break
            if isinstance(c.value,str) and c.value.strip(): label=c.value.strip()
        vals=[]; numeric=False
        for pc in per_cols:
            cell=ws.cell(row=row[0].row,column=pc)
            v=cell.value
            if isinstance(v,(int,float)): vals.append(v); numeric=True
            else: vals.append(None)
        if label and (numeric or label):
            lines.append({"label":label,"values":vals})
    return {"periods":periods,"lines":lines}

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("template")
    ap.add_argument("--out", default=None)
    a = ap.parse_args()

    wb_v = openpyxl.load_workbook(a.template, data_only=True, read_only=True)
    wb_f = openpyxl.load_workbook(a.template, data_only=False, read_only=True)

    result = {"file": a.template, "sheets": [], "segments": {}, "period_columns": {}, "cached_values_present": False}
    for name in wb_v.sheetnames:
        ws = wb_v[name]
        info = {"name": name, "max_row": ws.max_row, "max_col": ws.max_column}
        result["sheets"].append(info)
        if name.lower().startswith("segment") or name == "Segments":
            result["segments"] = find_segments_block(ws)
        low=name.lower()
        skey=None
        if "income" in low or low in ("is","p&l","pl"): skey="income_statement"
        elif "balance" in low or low=="bs": skey="balance_sheet"
        elif "cash" in low or low=="cf": skey="cash_flow"
        elif "ratio" in low: skey="ratios"
        if skey:
            stmt=extract_statement(ws)
            if stmt: result.setdefault("statements_raw",{})[skey]=stmt
        ph = period_headers(ws)
        if ph:
            result["period_columns"][name] = ph

    # detect any numeric cached value
    for name in wb_v.sheetnames:
        ws = wb_v[name]
        cnt = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 200)):
            for c in row:
                if isinstance(c.value, (int, float)):
                    cnt += 1
        if cnt > 5:
            result["cached_values_present"] = True
            break

    out = a.out or (a.template.rsplit(".",1)[0] + "_structure.json")
    json.dump(result, open(out, "w"), ensure_ascii=False, indent=2)
    print(f"wrote {out}")
    print(f"sheets: {[s['name'] for s in result['sheets']]}")
    print(f"segments: {result['segments'].get('meta', {})}")
    print(f"cached_values_present: {result['cached_values_present']}")

if __name__ == "__main__":
    main()
